"""Render the processed dashboard into a formatted .xlsx workbook.

The workbook is organized by warehouse. Three combined sheets come first
(All Plants Summary / Issue Tracker / Not Started), then one section per
warehouse detected in the export — its Issue Tracker, Not Started, Orders and
whichever customer reports actually have records — then the full SAPUI5 Export
and, when needed, Validation Warnings. Nothing is hardcoded to a fixed set of
warehouses: the sections come from ``result.site_sections``.

Formatting follows the spec: dark-blue headers, frozen header row, Excel
auto-filters, MM/DD/YYYY dates, Canadian-currency totals, whole-percent
picking, wrapped long text, status/priority highlighting and totals rows on the
customer detail reports.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import risk_engine as R
from .sheet_names import SheetNamer, site_label_map

# --- styles ------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="9C0006")

FILL_CRITICAL = PatternFill("solid", fgColor="FFC7CE")   # red
FILL_AT_RISK = PatternFill("solid", fgColor="FFD9A0")     # orange
FILL_NOT_STARTED = PatternFill("solid", fgColor="FFF2CC")  # yellow
FILL_READY = PatternFill("solid", fgColor="D9EAD3")       # green-ish

CURRENCY_FMT = '"$"#,##0'
DATE_FMT = "MM/DD/YYYY"
PCT_FMT = '0"%"'

_CURRENCY_COLS = {
    "Total Price", "Sales Order Total",
    "Total Order Value",                                         # Total orders block
    "Late Value", "Completed $", "Partial $", "Not Started $",   # Late orders — picking
}
_DATE_COLS = {
    "Warehouse Task Creat", "Route Depart. Date", "Planned Dlv. Date",
    "Customer Dock Appointment Date", "Due Date",
}
_PCT_COLS = {"Picking in %", "picking_pct"}
_WRAP_COLS = {"Issue", "Consequence / Risk", "Latest Update"}

# Whole-number columns; listed so a blank cell writes 0 rather than an empty
# string that Excel would refuse to sum.
_COUNT_COLS = {
    "Cases", "Line Items", "Pallet Quantity", "Num. Lifts Pallets", "Gross Weight",
    "Order Count", "Days to Route Departure", "Days to Planned Delivery",
    "Total Orders", "Open Orders", "Critical", "At Risk", "Not Started", "Late",
    "Route Departure Missed", "Issues",
    "Late Orders", "Completed", "Partial",          # Late orders — picking block
}

AMZ_NOTE = ("HIGHLIGHTED ORDERS ARE CURRENTLY LATE AND MAY BE SUBJECT TO THE "
            "CONFIGURED ON-TIME DELIVERY CHARGEBACK")

# The combined sheets, which are present in every download regardless of which
# warehouses the export contains. The per-warehouse sections are generated
# between "All Plants Not Started" and "SAPUI5 Export".
SHEET_ORDER = [
    "All Plants Summary", "All Plants Issue Tracker", "All Plants Not Started",
    "SAPUI5 Export",
]

# Written last, and only when the upload produced something to report.
VALIDATION_SHEET = "Validation Warnings"

# Suffixes appended to a warehouse label, in section order.
SITE_SHEET_SUFFIXES = ("Issue Tracker", "Not Started", "Orders")


def _sanitize(value, is_numeric_col: bool):
    """Convert a pandas cell into something openpyxl can write cleanly."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return 0 if is_numeric_col else None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value.to_pydatetime()
    try:
        if pd.isna(value):
            return 0 if is_numeric_col else None
    except (TypeError, ValueError):
        pass
    return value


def _fmt_for(col: str) -> str | None:
    """Number format for a column, by name. None = leave as General."""
    if col in _CURRENCY_COLS:
        return CURRENCY_FMT
    if col in _DATE_COLS:
        return DATE_FMT
    if col in _PCT_COLS:
        return PCT_FMT
    return None


def _numeric_cols(columns) -> set:
    """Columns whose blanks should be written as 0 rather than an empty cell."""
    return {c for c in columns
            if c in _CURRENCY_COLS or c in _PCT_COLS or c in _COUNT_COLS}


def _write_block(ws, title: str, df: pd.DataFrame, start_row: int) -> int:
    """Write a titled sub-table under the main sheet table.

    Returns the next free row (one blank row is left after the block).
    """
    ws.cell(row=start_row, column=1, value=title).font = TOTAL_FONT
    r = start_row + 1
    if df is None or df.empty:
        ws.cell(row=r, column=1, value="No data")
        return r + 2

    columns = list(df.columns)
    numeric = _numeric_cols(columns)
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=r, column=i, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    r += 1

    records = df.to_dict("records")
    for rec in records:
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=i, value=_sanitize(rec.get(col), col in numeric))
            fmt = _fmt_for(col)
            if fmt:
                cell.number_format = fmt
        r += 1

    # The roll-up sits on the last row of a multi-warehouse block; bold it the
    # same way the main summary table bolds its All Plants row.
    if len(records) > 1 and "Site" in columns:
        for i in range(1, len(columns) + 1):
            ws.cell(row=r - 1, column=i).font = TOTAL_FONT

    # Widen only where this block needs more room than the table above.
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        current = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = min(max(current, len(str(col)) + 4, 14), 50)
    return r + 1


def _detail_highlight(row: dict, as_of_ts: pd.Timestamp):
    """Row fill for a customer detail sheet, from raw canonical columns."""
    gi = str(row.get("Goods Issue", "") or "").strip().upper()
    planned = pd.to_datetime(row.get("Planned Dlv. Date"), errors="coerce")
    route = pd.to_datetime(row.get("Route Depart. Date"), errors="coerce")
    pct = pd.to_numeric(pd.Series([row.get("Picking in %")]), errors="coerce").fillna(0).iloc[0]
    late = pd.notna(planned) and planned < as_of_ts and gi != "COMPLETED"
    route_missed = pd.notna(route) and route < as_of_ts and pct < 100
    if late or route_missed:
        return FILL_CRITICAL
    if pct == 0 and gi != "COMPLETED":
        return FILL_NOT_STARTED
    return None


def _write_sheet(
    wb: Workbook,
    name: str,
    df: pd.DataFrame,
    *,
    as_of_ts: pd.Timestamp,
    add_totals: bool = False,
    note: str | None = None,
    status_col: str | None = None,
    priority_col: str | None = None,
    highlight_detail: bool = False,
    bold_last_row: bool = False,
    extra_blocks: list | None = None,
) -> None:
    # `name` is expected to already be legal and unique (see SheetNamer) —
    # openpyxl would silently rename a clash, which would hide the bug.
    ws = wb.create_sheet(title=name)
    if df is None or df.empty:
        ws.append([f"No data for: {name}"])
        _write_extra_blocks(ws, extra_blocks)
        return

    columns = list(df.columns)
    numeric_cols = _numeric_cols(columns)

    ws.append(columns)
    records = df.to_dict("records")
    for rec in records:
        ws.append([_sanitize(rec.get(c), c in numeric_cols) for c in columns])

    # Header styling.
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"

    # Column formats + widths.
    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        fmt = _fmt_for(col)
        if fmt:
            for r in range(2, len(records) + 2):
                ws.cell(row=r, column=idx).number_format = fmt
        if col in _WRAP_COLS:
            for r in range(2, len(records) + 2):
                ws.cell(row=r, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[letter].width = 46
        else:
            sample = [str(rec.get(col, "")) for rec in records[:200]]
            width = max([len(str(col))] + [len(s) for s in sample]) + 2
            ws.column_dimensions[letter].width = min(max(width, 12), 50)

    # Row highlighting.
    for r_i, rec in enumerate(records, start=2):
        fill = None
        if status_col and status_col in rec:
            s = rec[status_col]
            fill = {R.CRITICAL: FILL_CRITICAL, R.AT_RISK: FILL_AT_RISK,
                    R.NOT_STARTED: FILL_NOT_STARTED, R.READY: FILL_READY}.get(s)
        elif priority_col and priority_col in rec:
            p = rec[priority_col]
            fill = {"Critical": FILL_CRITICAL, "High": FILL_AT_RISK,
                    "Medium": FILL_NOT_STARTED}.get(p)
        elif highlight_detail:
            fill = _detail_highlight(rec, as_of_ts)
        if fill:
            for c_i in range(1, len(columns) + 1):
                ws.cell(row=r_i, column=c_i).fill = fill

    # Roll-up row of the All Plants Summary.
    if bold_last_row:
        for c_i in range(1, len(columns) + 1):
            ws.cell(row=len(records) + 1, column=c_i).font = TOTAL_FONT

    # Totals row for customer reports.
    if add_totals and "Sales Order Total" in columns:
        total = float(pd.to_numeric(df["Sales Order Total"], errors="coerce").fillna(0).sum())
        total_row = len(records) + 2
        label_col = 1
        ws.cell(row=total_row, column=label_col, value="TOTAL").font = TOTAL_FONT
        tcol = columns.index("Sales Order Total") + 1
        cell = ws.cell(row=total_row, column=tcol, value=total)
        cell.number_format = CURRENCY_FMT
        cell.font = TOTAL_FONT

    if note:
        note_row = ws.max_row + 2
        c = ws.cell(row=note_row, column=1, value=note)
        c.font = NOTE_FONT

    _write_extra_blocks(ws, extra_blocks)


def _write_extra_blocks(ws, extra_blocks: list | None) -> None:
    """Write each ``(title, dataframe)`` block below whatever is on the sheet."""
    for title, block in extra_blocks or []:
        _write_block(ws, title, block, ws.max_row + 2)


def sheet_plan(result) -> list[tuple[str, object, dict]]:
    """``(requested name, dataframe, formatting kwargs)`` for every sheet, in order.

    Names here are the *intended* ones; :class:`SheetNamer` in
    :func:`build_workbook` is what turns them into legal, unique worksheet
    titles. Split out from the writing so the plan can be asserted in tests
    without building a workbook.
    """
    plan: list[tuple[str, object, dict]] = [
        ("All Plants Summary", result.all_plants_summary,
         {"bold_last_row": True,
          "extra_blocks": list(getattr(result, "summary_blocks", []) or [])}),
        ("All Plants Issue Tracker", result.issue_tracker, {"status_col": "Status"}),
        ("All Plants Not Started", result.not_started, {"priority_col": "Priority"}),
    ]

    sections = list(getattr(result, "site_sections", []) or [])
    labels = site_label_map(s.site for s in sections)
    for section in sections:
        label = labels[section.site]
        plan.append((f"{label} Issue Tracker", section.issue_tracker, {"status_col": "Status"}))
        plan.append((f"{label} Not Started", section.not_started, {"priority_col": "Priority"}))
        plan.append((f"{label} Orders", section.orders,
                     {"add_totals": True, "highlight_detail": True}))
        # Customer reports for this warehouse; build_site_sections has already
        # dropped the ones with no matching records, so none of these are empty.
        for name, df in section.reports.items():
            attrs = getattr(df, "attrs", None) or {}
            is_customer_report = bool(attrs.get("is_customer_report"))
            plan.append((f"{label} - {name}", df, {
                "add_totals": is_customer_report,
                "highlight_detail": is_customer_report,
                "note": AMZ_NOTE if attrs.get("amazon_note") else None,
            }))

    plan.append(("SAPUI5 Export", (result.reports or {}).get("SAPUI5 Export"), {}))

    errors = result.validation_errors
    if errors is not None and not errors.empty:
        plan.append((VALIDATION_SHEET, errors, {}))
    return plan


def build_workbook(result) -> bytes:
    """Build the full workbook from a ``ProcessResult``. Returns .xlsx bytes."""
    as_of_ts = pd.Timestamp(result.as_of)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    namer = SheetNamer()
    for name, df, options in sheet_plan(result):
        _write_sheet(wb, namer.allocate(name), df, as_of_ts=as_of_ts, **options)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
