"""Tests for the Daily Delivery Risk Dashboard engine.

Covers header normalization, Excel date parsing, identifier cleaning, picking
conversion, LE Delivery de-duplication, customer classification, Not Started /
late logic, Total Price, Amazon deadline logic, Andrew Tab filtering, Excel
report creation, and the "new upload replaces old data" guarantee.

The prototype workbook (Calgary_July_17) is used for end-to-end validation when
present; those checks skip gracefully in environments without the file so the
core unit tests always run.
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from delivery_dashboard import cleaner, loader, process_export, risk_engine
from delivery_dashboard.customer_rules import load_ruleset
from delivery_dashboard.excel_exporter import (
    SHEET_ORDER,
    VALIDATION_SHEET,
    build_workbook,
    sheet_plan,
)
from delivery_dashboard.loader import CANONICAL_COLUMNS, DeliveryLoadError
from delivery_dashboard.report_builder import (
    ALL_PLANTS,
    ISSUE_TRACKER_COLUMNS,
    build_andrew_tab,
    build_andrew_tab_2,
    build_all_plants_summary,
    build_issue_tracker,
    build_not_started,
)
from delivery_dashboard.sheet_names import (
    MAX_SHEET_NAME,
    SheetNamer,
    sanitize_sheet_name,
    shorten,
    site_label_map,
)

PROTOTYPE = Path(r"C:/Users/melgh/Downloads/Calgary_July_17 (Prototype).xlsx")
AS_OF = date(2026, 7, 17)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _xlsx(headers: list, rows: list[list], sheet="SAPUI5 Export") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# A minimal but complete synthetic export builder — one row = kwargs override.
_DEFAULTS = {
    "ys": "Calgary Warehouse", "Facility Code": "Warehouse", "Warehouse Task Creat": "Y",
    "Picking in %": "0", "Route Depart. Date": "2026-07-20 23:00:00",
    "Planned Dlv. Date": "2026-07-22 23:00:00", "LE Delivery": "80100000",
    "Ship-to Name": "SOME CUSTOMER", "Carrier Description": "VERSACOLD LOGISTICS",
    "Cases": "10", "Line Items": "2", "Goods Issue": "Not Started",
    "Sales Order": "3000000", "Sales Order Total": "1000",
    "Customer Dock Appointment Date": None,
}


def _rows(*overrides: dict) -> io.BytesIO:
    headers = list(_DEFAULTS.keys())
    rows = []
    for i, ov in enumerate(overrides):
        rec = dict(_DEFAULTS)
        rec["LE Delivery"] = str(80100000 + i)
        rec.update(ov)
        rows.append([rec[h] for h in headers])
    return _xlsx(headers, rows)


def _process(buf, as_of=AS_OF):
    return process_export(buf, as_of)


def _multi_site() -> io.BytesIO:
    """Two warehouses, each with its own Amazon problem plus a second customer."""
    return _rows(
        {"ys": "Calgary Warehouse", "Ship-to Name": "AMAZON CANADA",
         "Carrier Description": "AMAZON PICKUP",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "5000"},
        {"ys": "Calgary Warehouse", "Ship-to Name": "GFS CALGARY",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "2000"},
        {"ys": "Mississauga Warehouse", "Ship-to Name": "AMAZON ONTARIO",
         "Carrier Description": "AMAZON PICKUP",
         "Planned Dlv. Date": "2026-07-11 23:00:00", "Sales Order Total": "7500"},
        {"ys": "Mississauga Warehouse", "Ship-to Name": "SYSCO TORONTO",
         "Planned Dlv. Date": "2026-07-11 23:00:00", "Sales Order Total": "3000"},
    )


# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------
def test_header_normalization_tolerates_spacing_case_newlines():
    assert loader.resolve_column("  ship-to name ") == "Ship-to Name"
    assert loader.resolve_column("SHIP-TO NAME") == "Ship-to Name"
    assert loader.resolve_column("Planned Dlv. Date\n") == "Planned Dlv. Date"
    assert loader.resolve_column("Picking in %") == "Picking in %"
    assert loader.resolve_column("Route Departure Date") == "Route Depart. Date"
    assert loader.resolve_column("totally unknown column") is None


def test_missing_required_columns_raise():
    buf = _xlsx(["ys", "LE Delivery", "Cases"], [["Calgary", "1", "3"]])
    with pytest.raises(DeliveryLoadError):
        loader.load_sap_export(buf)


def test_empty_workbook_raises():
    buf = _xlsx(list(_DEFAULTS.keys()), [])
    with pytest.raises(DeliveryLoadError):
        loader.load_sap_export(buf)


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------
def test_identifier_cleaning_strips_dot_zero_and_keeps_string():
    df = loader.load_sap_export(_rows({"LE Delivery": "80117458.0", "Sales Order": "3075382.0"}))
    clean, _ = cleaner.clean(df)
    assert clean["LE Delivery"].iloc[0] == "80117458"
    assert clean["Sales Order"].iloc[0] == "3075382"


def test_excel_serial_date_parsing():
    # 46220 == 2026-07-17 in the Excel 1900 date system (epoch 1899-12-30).
    df = loader.load_sap_export(_rows({"Planned Dlv. Date": "46220"}))
    clean, _ = cleaner.clean(df)
    assert pd.Timestamp(clean["Planned Dlv. Date"].iloc[0]).date() == date(2026, 7, 17)


def test_datetime_string_date_parsing():
    df = loader.load_sap_export(_rows({"Planned Dlv. Date": "2026-07-22 23:00:00"}))
    clean, _ = cleaner.clean(df)
    assert pd.Timestamp(clean["Planned Dlv. Date"].iloc[0]).date() == date(2026, 7, 22)


def test_picking_percentage_conversion():
    df = loader.load_sap_export(_rows(
        {"Picking in %": "70"}, {"Picking in %": "not-a-number"}, {"Picking in %": "150"}))
    clean, _ = cleaner.clean(df)
    pct = clean.sort_values("LE Delivery")["Picking in %"].tolist()
    assert pct[0] == 70.0
    assert pct[1] == 0.0        # invalid -> 0
    assert pct[2] == 100.0      # clipped


def test_picking_fraction_column_is_scaled():
    df = loader.load_sap_export(_rows(
        {"Picking in %": "0.7"}, {"Picking in %": "1"}, {"Picking in %": "0"}))
    clean, _ = cleaner.clean(df)
    assert clean["Picking in %"].max() == 100.0  # 0..1 column scaled to 0..100


def test_bad_sales_order_total_becomes_zero_and_is_flagged():
    # "TBD" is non-numeric junk that pandas does NOT auto-treat as NA.
    df = loader.load_sap_export(_rows({"Sales Order Total": "TBD"}))
    clean, errors = cleaner.clean(df)
    assert clean["Sales Order Total"].iloc[0] == 0.0
    assert (errors["Validation Issue"].str.contains("Sales Order Total")).any()


def test_blank_rows_removed_and_missing_delivery_quarantined():
    df = loader.load_sap_export(_rows(
        {"LE Delivery": "80100001"},
        {"LE Delivery": ""},   # no delivery -> quarantined
    ))
    clean, errors = cleaner.clean(df)
    assert len(clean) == 1
    assert (errors["Validation Issue"].str.contains("Missing LE Delivery")).any()


def test_dedup_by_le_delivery_no_double_count():
    df = loader.load_sap_export(_rows(
        {"LE Delivery": "80100001", "Sales Order Total": "1000", "Picking in %": "0"},
        {"LE Delivery": "80100001", "Sales Order Total": "1000", "Picking in %": "50"},
    ))
    clean, _ = cleaner.clean(df)
    assert len(clean) == 1
    assert clean["Sales Order Total"].sum() == 1000.0


# ---------------------------------------------------------------------------
# Customer classification
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("ship,carrier,expected", [
    ("AMAZON CANADA FULFILLMENT", "VERSACOLD", "amazon"),
    ("SOME STORE", "AMAZON PICKUP", "amazon"),
    ("AMZ WAREHOUSE 3", "VERSACOLD", "amazon"),
    ("GFS EDMONTON", "VERSACOLD", "gfs"),
    ("A&W GFS CALGARY", "GFS CAPSTONE PICKUP", "aw_via_gfs"),
    ("SYSCO CALGARY", "VERSACOLD", "sysco"),
    ("CALGARY CO-OP 704", "VERSACOLD", "calgary_coop"),
    ("SAVE ON C/O TCL SUPPLY", "SAVE ON PICKUP", "save_on"),
    ("ASSOCIATED GROCERS 9221 WHS", "VERSACOLD", "associated_grocers"),
    ("FCL 4840 WAREHOUSE", "VERSACOLD", "federated_coop"),
    ("SOBEYS 3012 MILLWOODS", "VERSACOLD", "sobeys_group"),
    ("CANADA SAFEWAY 8830", "VERSACOLD", "sobeys_group"),
    ("PRATT'S WHOLESALE LTD", "VERIGO", "pratts"),
    ("UNKNOWN CORNER STORE", "SCOOPS", "other"),
])
def test_customer_classification(ship, carrier, expected):
    rules = load_ruleset()
    df = loader.load_sap_export(_rows({"Ship-to Name": ship, "Carrier Description": carrier}))
    clean, _ = cleaner.clean(df)
    classified = rules.classify(clean)
    assert classified["customer_group"].iloc[0] == expected


@pytest.mark.parametrize("ship, carrier, expected", [
    # Walmart
    ("SOME RECEIVER", "WALMART PICKUP", "walmart"),
    ("SOME RECEIVER", "WALMART PICK UP", "walmart"),
    ("WALMART DC 6094", "VERSACOLD", "walmart"),
    ("WM CALGARY DC", "VERSACOLD", "walmart"),
    # "WM " is a prefix rule, not a substring: neither of these is Walmart.
    ("WMX LOGISTICS", "VERSACOLD", "other"),
    ("BUILDING WM SUPPLY", "VERSACOLD", "other"),
    # Loblaws
    ("SOME RECEIVER", "LOBLAWS PICK UP", "loblaws"),
    ("SOME RECEIVER", "LOBLAW PICK UP", "loblaws"),
    ("SOME RECEIVER", "LOBLAWS PICKUP", "loblaws"),
    ("SOME RECEIVER", "LOBLAW PICKUP", "loblaws"),
    ("LOBLAW COMPANIES LTD", "VERSACOLD", "loblaws"),
    ("LOBLAWS SUPERMARKET 1234", "VERSACOLD", "loblaws"),
    # Consignee names only count as Loblaws when the carrier says so.
    ("WESTERN GROCERS", "LOBLAWS PICK UP", "loblaws"),
    ("WESTERN GROCERS", "VERSACOLD", "other"),
    ("ATLANTIC FREEZER DIST CENTRE", "LOBLAW PICKUP", "loblaws"),
    ("ATLANTIC FREEZER DIST CENTRE", "VERSACOLD", "other"),
])
def test_walmart_and_loblaws_classification(ship, carrier, expected):
    rules = load_ruleset()
    df = loader.load_sap_export(_rows({"Ship-to Name": ship, "Carrier Description": carrier}))
    clean, _ = cleaner.clean(df)
    classified = rules.classify(clean)
    assert classified["customer_group"].iloc[0] == expected


def test_walmart_and_loblaws_display_names():
    rules = load_ruleset()
    assert rules.rule("walmart").display == "Walmart"
    assert rules.rule("loblaws").display == "Loblaws"


def test_existing_customers_still_win_over_the_new_rules():
    # Every rule above Walmart/Loblaws keeps its match — the new groups only
    # take orders that were falling through to the fallback.
    rules = load_ruleset()
    df = loader.load_sap_export(_rows(
        {"Ship-to Name": "AMAZON YYC4", "Carrier Description": "WALMART PICKUP"},
        {"Ship-to Name": "SOBEYS 3012", "Carrier Description": "LOBLAWS PICKUP"},
    ))
    clean, _ = cleaner.clean(df)
    groups = rules.classify(clean).sort_values("LE Delivery")["customer_group"].tolist()
    assert groups == ["amazon", "sobeys_group"]


def test_new_mappings_only_move_orders_out_of_other():
    """Deliveries, totals and every other classification are untouched."""
    buf_rows = [
        {"Ship-to Name": "WALMART DC 6094", "Sales Order Total": "1500"},
        {"Ship-to Name": "WESTERN GROCERS", "Carrier Description": "LOBLAWS PICK UP",
         "Sales Order Total": "2500"},
        {"Ship-to Name": "WESTERN GROCERS", "Sales Order Total": "3500"},
        {"Ship-to Name": "SOBEYS 3012", "Sales Order Total": "4500"},
    ]
    result = _process(_rows(*buf_rows))
    orders = result.orders.sort_values("LE Delivery")

    assert orders["customer_group"].tolist() == [
        "walmart", "loblaws", "other", "sobeys_group"]
    assert orders["customer_display"].tolist() == [
        "Walmart", "Loblaws", "Other / Unclassified", "Sobeys/Safeway/FreshCo"]
    # No delivery gained or lost, no value gained or lost.
    assert orders["LE Delivery"].nunique() == len(buf_rows) == len(orders)
    assert orders["Sales Order Total"].sum() == 12000.0


def test_new_mappings_do_not_change_risk_status():
    """Walmart/Loblaws keep the fallback priority, so no order changes status."""
    rows = [
        {"Ship-to Name": "WALMART DC 6094"},
        {"Ship-to Name": "SOME RECEIVER", "Carrier Description": "LOBLAWS PICKUP"},
        {"Ship-to Name": "UNKNOWN CORNER STORE"},
    ]
    statuses = _process(_rows(*rows)).orders.sort_values("LE Delivery")["risk_status"]
    # All three are the same order shape, so all three must land on one status.
    assert statuses.nunique() == 1


def test_walmart_and_loblaws_group_on_the_existing_issue_tracker():
    result = _process(_rows(
        # Late -> escalating, so both appear on the tracker as their own rows.
        {"ys": "Calgary Warehouse", "Ship-to Name": "WALMART DC 6094",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "1500"},
        {"ys": "Calgary Warehouse", "Ship-to Name": "WESTERN GROCERS",
         "Carrier Description": "LOBLAWS PICK UP",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "2500"},
    ))
    tracker = result.issue_tracker
    assert list(tracker.columns) == ISSUE_TRACKER_COLUMNS      # unchanged template
    assert set(tracker["Customer"]) == {"Walmart", "Loblaws"}
    assert "Other / Unclassified" not in set(tracker["Customer"])
    assert tracker["Order Count"].sum() == 2
    assert tracker["Total Price"].sum() == 4000.0


def test_aw_matches_before_gfs():
    # A&W must win over the generic GFS group (listed first in config).
    rules = load_ruleset()
    df = loader.load_sap_export(_rows({"Ship-to Name": "A&W GFS CALGARY"}))
    clean, _ = cleaner.clean(df)
    assert rules.classify(clean)["customer_group"].iloc[0] == "aw_via_gfs"


# ---------------------------------------------------------------------------
# Risk / Not Started / late logic
# ---------------------------------------------------------------------------
def test_not_started_logic():
    df = loader.load_sap_export(_rows(
        {"Picking in %": "0", "Goods Issue": "Not Started"},      # not started
        {"Picking in %": "0", "Goods Issue": "Completed"},        # GI complete -> not
        {"Picking in %": "30", "Goods Issue": "Not Started"},     # in progress -> not
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert enriched.sort_values("LE Delivery")["is_not_started"].tolist() == [True, False, False]


def test_late_logic():
    df = loader.load_sap_export(_rows(
        {"Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Not Started"},  # late
        {"Planned Dlv. Date": "2026-07-25 23:00:00", "Goods Issue": "Not Started"},  # future
        {"Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Completed"},    # done -> not late
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert enriched.sort_values("LE Delivery")["is_late"].tolist() == [True, False, False]


def test_route_departure_missed_and_status_critical():
    df = loader.load_sap_export(_rows(
        {"Route Depart. Date": "2026-07-10 23:00:00", "Picking in %": "40",
         "Planned Dlv. Date": "2026-07-25 23:00:00"},
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert bool(enriched["is_route_departure_missed"].iloc[0]) is True
    assert enriched["risk_status"].iloc[0] == risk_engine.CRITICAL


def test_ready_awaiting_gi_status():
    df = loader.load_sap_export(_rows(
        {"Picking in %": "100", "Goods Issue": "Not Started",
         "Planned Dlv. Date": "2026-07-25 23:00:00", "Route Depart. Date": "2026-07-24 23:00:00"},
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert enriched["risk_status"].iloc[0] == risk_engine.READY


# ---------------------------------------------------------------------------
# Amazon deadline logic
# ---------------------------------------------------------------------------
def test_amazon_deadline_passed_is_critical():
    df = loader.load_sap_export(_rows(
        {"Ship-to Name": "AMAZON CANADA", "Carrier Description": "AMAZON PICKUP",
         "Customer Dock Appointment Date": "2026-07-10 00:00:00",
         "Planned Dlv. Date": "2026-07-25 23:00:00", "Route Depart. Date": "2026-07-24 23:00:00",
         "Picking in %": "50"},
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert enriched["risk_status"].iloc[0] == risk_engine.CRITICAL
    assert enriched["days_to_customer_deadline"].iloc[0] < 0


def test_amazon_deadline_within_five_days_is_at_risk():
    df = loader.load_sap_export(_rows(
        {"Ship-to Name": "AMAZON CANADA", "Carrier Description": "AMAZON PICKUP",
         "Customer Dock Appointment Date": "2026-07-21 00:00:00",  # 4 days out
         "Planned Dlv. Date": "2026-07-28 23:00:00", "Route Depart. Date": "2026-07-27 23:00:00",
         "Picking in %": "80"},
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    assert enriched["risk_status"].iloc[0] == risk_engine.AT_RISK


# ---------------------------------------------------------------------------
# Andrew Tab filtering
# ---------------------------------------------------------------------------
def test_andrew_tab_filter():
    df = loader.load_sap_export(_rows(
        {"Facility Code": "Warehouse", "Planned Dlv. Date": "2026-07-10 23:00:00", "Picking in %": "40"},   # in
        {"Facility Code": "Warehouse", "Planned Dlv. Date": "2026-07-10 23:00:00", "Picking in %": "100"},  # out (100%)
        {"Facility Code": "DSD", "Planned Dlv. Date": "2026-07-10 23:00:00", "Picking in %": "40"},         # out (facility)
        {"Facility Code": "Warehouse", "Planned Dlv. Date": "2026-07-25 23:00:00", "Picking in %": "40"},   # out (future)
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    tab = build_andrew_tab(enriched, AS_OF)
    assert len(tab) == 1
    assert list(tab.columns) == [
        "Picking in %", "Planned Dlv. Date", "LE Delivery", "Ship-to Name",
        "Ship-to City", "Ship-to Province", "Cases", "Line Items", "Sales Order Total"]


def test_andrew_tab_2_normalizes_spelling_and_filters():
    df = loader.load_sap_export(_rows(
        {"Facility Code": "Independant", "Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Not Started"},  # in
        {"Facility Code": "Independent", "Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Not Started"},  # in
        {"Facility Code": "DSD", "Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Not Started"},          # in
        {"Facility Code": "DSD", "Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Completed"},            # out
        {"Facility Code": "Warehouse", "Planned Dlv. Date": "2026-07-10 23:00:00", "Goods Issue": "Not Started"},    # out
    ))
    rules = load_ruleset()
    clean, _ = cleaner.clean(df)
    enriched = risk_engine.enrich(rules.classify(clean), AS_OF, rules)
    tab = build_andrew_tab_2(enriched, AS_OF)
    assert len(tab) == 3


# ---------------------------------------------------------------------------
# Total Price (issue tracker sums only affected orders, dedup by LE Delivery)
# ---------------------------------------------------------------------------
def test_issue_tracker_total_price_counts_only_escalating():
    buf = _rows(
        {"Ship-to Name": "SYSCO CALGARY", "Planned Dlv. Date": "2026-07-10 23:00:00",
         "Goods Issue": "Not Started", "Sales Order Total": "500"},                      # late -> escalating
        {"Ship-to Name": "SYSCO EDMONTON", "Planned Dlv. Date": "2026-08-30 23:00:00",
         "Route Depart. Date": "2026-08-29 23:00:00", "Goods Issue": "Not Started",
         "Sales Order Total": "999"},                                                    # far future -> not escalating
    )
    result = _process(buf)
    sysco_row = result.issue_tracker[result.issue_tracker["Customer"] == "Sysco"]
    assert len(sysco_row) == 1
    assert sysco_row["Total Price"].iloc[0] == 500.0   # 999 order excluded
    assert sysco_row["Order Count"].iloc[0] == 1


# ---------------------------------------------------------------------------
# Excel report creation
# ---------------------------------------------------------------------------
def test_excel_workbook_has_sheets_in_order():
    df = _rows({"Ship-to Name": "AMAZON CANADA", "Carrier Description": "AMAZON PICKUP",
                "Planned Dlv. Date": "2026-07-10 23:00:00"})
    result = _process(df)
    xlsx = build_workbook(result)
    wb = load_workbook(io.BytesIO(xlsx))
    for name in SHEET_ORDER:
        assert name in wb.sheetnames
    # Order of the canonical sheets is preserved.
    positions = [wb.sheetnames.index(n) for n in SHEET_ORDER]
    assert positions == sorted(positions)


def test_new_upload_replaces_old_data():
    """Processing a second file must not retain anything from the first."""
    first = _process(_rows({"Ship-to Name": "SYSCO CALGARY",
                            "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "500"}))
    assert "Sysco" in first.issue_tracker["Customer"].values

    second = _process(_rows({"Ship-to Name": "PRATT'S WHOLESALE",
                             "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "800"}))
    # No Sysco carried over; only Pratts present.
    assert "Sysco" not in second.issue_tracker["Customer"].values
    assert "Pratts" in second.issue_tracker["Customer"].values
    assert len(second.orders) == 1


# ---------------------------------------------------------------------------
# Warehouse label resolution
# ---------------------------------------------------------------------------
def test_site_label_treats_every_pandas_null_as_unknown():
    rules = load_ruleset()
    # pd.NA / NaT are not floats, so a naive isinstance check let the literal
    # text "<NA>" through and it showed up as a warehouse name.
    for missing in (None, pd.NA, pd.NaT, float("nan"), "", "   "):
        assert rules.site_label(missing) == "Unknown", repr(missing)
    assert rules.site_label("Calgary Warehouse") == "Calgary"


def test_missing_site_column_gives_one_unknown_warehouse_and_a_warning():
    headers = [h for h in _DEFAULTS if h != "ys"]
    rec = dict(_DEFAULTS)
    result = _process(_xlsx(headers, [[rec[h] for h in headers]]))
    assert result.sites == ["Unknown"]
    assert any("warehouse column" in w for w in result.warnings)


@pytest.mark.parametrize("header", ["ys", "Site", "Warehouse", "Plant"])
def test_site_column_is_recognized_under_its_common_names(header):
    headers = [header] + [h for h in _DEFAULTS if h != "ys"]
    rec = dict(_DEFAULTS)
    rec[header] = "Calgary Warehouse"
    result = _process(_xlsx(headers, [[rec[h] for h in headers]]))
    assert result.sites == ["Calgary"]
    assert result.warnings == []


# ---------------------------------------------------------------------------
# Safe sheet names
# ---------------------------------------------------------------------------
def test_sanitize_removes_invalid_characters():
    assert sanitize_sheet_name("Save/On:Foods*?[x]") == "Save On Foods x"
    assert sanitize_sheet_name("'Calgary'") == "Calgary"
    assert sanitize_sheet_name("///") == "Sheet"        # never empty
    assert sanitize_sheet_name(None) == "Sheet"


def test_shorten_prefers_a_word_boundary():
    assert shorten("Calgary North Distribution", 15) == "Calgary North"
    # A single word longer than the limit still has to be cut somewhere.
    assert shorten("Supercalifragilistic", 8) == "Supercal"


def test_sheet_namer_caps_length_and_deduplicates():
    namer = SheetNamer()
    assert namer.allocate("Calgary - AMZ") == "Calgary - AMZ"

    long_name = "Mississauga Distribution Centre - Calgary CO-OP Warehouse"
    first, second, third = (namer.allocate(long_name) for _ in range(3))
    assert first != second != third
    assert second.endswith("(2)") and third.endswith("(3)")
    for name in (first, second, third):
        assert len(name) <= MAX_SHEET_NAME


def test_sheet_namer_avoids_the_name_excel_reserves():
    assert SheetNamer().allocate("History") != "History"


def test_site_label_map_is_stable_and_leaves_room_for_a_suffix():
    assert site_label_map(["Calgary", "Mississauga"]) == {
        "Calgary": "Calgary", "Mississauga": "Mississauga"}
    long_site = "Calgary North Regional Distribution Centre"
    assert site_label_map([long_site])[long_site] == "Calgary North"


# ---------------------------------------------------------------------------
# Warehouse sections
# ---------------------------------------------------------------------------
def test_sites_are_detected_dynamically_and_ordered_by_config():
    buf = _rows(
        {"ys": "Surrey Warehouse"},
        {"ys": "Calgary Warehouse"},
        {"ys": "Mississauga Warehouse"},
        {"ys": "Winnipeg Warehouse"},      # not in site_order -> alphabetical, last
    )
    # Configured warehouses keep config order; an unconfigured one still gets a
    # section rather than being dropped.
    assert _process(buf).sites == ["Mississauga", "Calgary", "Surrey", "Winnipeg"]


def test_all_plants_tracker_keeps_the_same_customer_separate_per_warehouse():
    result = _process(_multi_site())
    assert "Site" in result.issue_tracker.columns
    amazon = result.issue_tracker[result.issue_tracker["Customer"] == "Amazon"]
    assert sorted(amazon["Site"]) == ["Calgary", "Mississauga"]


def test_all_plants_tracker_splits_one_customer_by_issue_type():
    buf = _rows(
        {"ys": "Calgary Warehouse", "Ship-to Name": "AMAZON CANADA",
         "Carrier Description": "AMAZON PICKUP",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Sales Order Total": "5000"},
        {"ys": "Calgary Warehouse", "Ship-to Name": "AMAZON CANADA",
         "Carrier Description": "AMAZON PICKUP",
         "Customer Dock Appointment Date": "2026-07-19 00:00:00",
         "Planned Dlv. Date": "2026-08-20 23:00:00", "Route Depart. Date": "2026-08-19 23:00:00",
         "Picking in %": "80", "Sales Order Total": "3000"},
    )
    amazon = _process(buf).issue_tracker
    amazon = amazon[amazon["Customer"] == "Amazon"]
    assert len(amazon) == 2
    assert set(amazon["Issue Type"]) == {risk_engine.IT_LATE, risk_engine.IT_DEADLINE_SOON}


def test_warehouse_tracker_is_exactly_the_all_plants_slice():
    result = _process(_multi_site())
    keys = ["Site", "Customer", "Issue Type", "Status", "Order Count", "Total Price"]
    for section in result.site_sections:
        expected = result.issue_tracker[result.issue_tracker["Site"] == section.site]
        pd.testing.assert_frame_equal(
            section.issue_tracker[keys].reset_index(drop=True),
            expected[keys].reset_index(drop=True),
        )


def test_warehouse_not_started_uses_picking_zero_and_normalized_goods_issue():
    buf = _rows(
        {"ys": "Calgary Warehouse", "Picking in %": "0", "Goods Issue": "Not Started"},   # in
        {"ys": "Calgary Warehouse", "Picking in %": "0", "Goods Issue": "COMPLETED"},     # out
        {"ys": "Calgary Warehouse", "Picking in %": "25", "Goods Issue": "Not Started"},  # out
        {"ys": "Surrey Warehouse", "Picking in %": "0", "Goods Issue": "Not Started"},    # other site
    )
    sections = {s.site: s for s in _process(buf).site_sections}
    assert len(sections["Calgary"].not_started) == 1
    assert set(sections["Calgary"].not_started["Site"]) == {"Calgary"}
    assert len(sections["Surrey"].not_started) == 1


def test_warehouse_orders_hold_every_unique_delivery_exactly_once():
    result = _process(_multi_site())
    assert sum(len(s.orders) for s in result.site_sections) == len(result.orders)
    ids = pd.concat([s.orders["LE Delivery"] for s in result.site_sections])
    assert ids.is_unique


def test_duplicate_delivery_is_not_double_counted_in_warehouse_figures():
    buf = _rows(
        {"ys": "Calgary Warehouse", "LE Delivery": "DUP1", "Sales Order Total": "1000",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Picking in %": "0"},
        {"ys": "Calgary Warehouse", "LE Delivery": "DUP1", "Sales Order Total": "1000",
         "Planned Dlv. Date": "2026-07-10 23:00:00", "Picking in %": "50"},
    )
    result = _process(buf)
    calgary = result.all_plants_summary[result.all_plants_summary["Site"] == "Calgary"].iloc[0]
    assert calgary["Total Orders"] == 1
    assert calgary["Total Price $"] == 1000.0
    assert result.issue_tracker["Order Count"].sum() == 1


def test_all_plants_summary_column_layout():
    summary = _process(_multi_site()).all_plants_summary
    assert list(summary.columns) == [
        "Site", "Total Orders", "Total Price $", "Late Orders", "Late Value",
        "Completed", "Completed $", "Partial", "Partial $",
        "Not Started", "Not Started $"]


def test_late_picking_buckets_partition_the_late_orders():
    buf = _rows(
        # Three late orders (planned date in the past, still open), one in each
        # picking bucket, plus a future-dated order that must not appear.
        {"ys": "Calgary Warehouse", "LE Delivery": "L1", "Sales Order Total": "1000",
         "Planned Dlv. Date": "2026-07-10 00:00:00", "Picking in %": "100"},
        {"ys": "Calgary Warehouse", "LE Delivery": "L2", "Sales Order Total": "2000",
         "Planned Dlv. Date": "2026-07-10 00:00:00", "Picking in %": "50"},
        {"ys": "Calgary Warehouse", "LE Delivery": "L3", "Sales Order Total": "4000",
         "Planned Dlv. Date": "2026-07-10 00:00:00", "Picking in %": "0"},
        {"ys": "Calgary Warehouse", "LE Delivery": "F1", "Sales Order Total": "9000",
         "Planned Dlv. Date": "2026-08-30 00:00:00", "Picking in %": "0"},
    )
    result = _process(buf)
    row = result.all_plants_summary.iloc[0]

    assert row["Total Orders"] == 4
    assert row["Total Price $"] == 16000.0
    assert row["Late Orders"] == 3
    assert row["Late Value"] == 7000.0
    assert (row["Completed"], row["Completed $"]) == (1, 1000.0)
    assert (row["Partial"], row["Partial $"]) == (1, 2000.0)
    assert (row["Not Started"], row["Not Started $"]) == (1, 4000.0)
    # The buckets partition the late orders exactly.
    assert row["Completed"] + row["Partial"] + row["Not Started"] == row["Late Orders"]
    assert (row["Completed $"] + row["Partial $"] + row["Not Started $"]
            == row["Late Value"])


def test_summary_sheet_is_one_table_with_no_extra_blocks():
    result = _process(_multi_site())
    wb = load_workbook(io.BytesIO(build_workbook(result)))
    ws = wb["All Plants Summary"]

    # Header row, one row per warehouse, one roll-up row — nothing below it.
    assert ws.max_row == len(result.sites) + 2
    assert [c.value for c in ws[1]] == list(result.all_plants_summary.columns)
    assert ws.cell(row=ws.max_row, column=1).value == ALL_PLANTS


def test_customer_reports_are_only_created_where_records_exist():
    sections = {s.site: s for s in _process(_multi_site()).site_sections}
    assert all(not df.empty for s in sections.values() for df in s.reports.values())
    # Only Mississauga has a Sysco order in this export.
    assert "Sysco" in sections["Mississauga"].reports
    assert "Sysco" not in sections["Calgary"].reports
    assert "GFS" in sections["Calgary"].reports
    assert "GFS" not in sections["Mississauga"].reports


def test_all_plants_summary_rolls_up_every_warehouse():
    result = _process(_multi_site())
    summary = result.all_plants_summary
    assert list(summary["Site"]) == result.sites + [ALL_PLANTS]

    total = summary[summary["Site"] == ALL_PLANTS].iloc[0]
    per_site = summary[summary["Site"] != ALL_PLANTS]
    assert total["Total Orders"] == per_site["Total Orders"].sum() == len(result.orders)
    for col in ["Total Price $", "Late Orders", "Late Value",
                "Completed", "Completed $", "Partial", "Partial $",
                "Not Started", "Not Started $"]:
        assert total[col] == per_site[col].sum(), col


# ---------------------------------------------------------------------------
# Warehouse-organized workbook
# ---------------------------------------------------------------------------
def test_workbook_sheet_order_follows_the_warehouse_sections():
    result = _process(_multi_site())
    names = [name for name, _df, _options in sheet_plan(result)]

    assert names[:3] == [
        "All Plants Summary", "All Plants Issue Tracker", "All Plants Not Started"]
    assert names[-1] == "SAPUI5 Export"

    assert result.sites == ["Mississauga", "Calgary"]
    for site in result.sites:
        start = names.index(f"{site} Issue Tracker")
        assert names[start:start + 3] == [
            f"{site} Issue Tracker", f"{site} Not Started", f"{site} Orders"]
    assert names.index("Mississauga Orders") < names.index("Calgary Issue Tracker")

    # Everything between the combined sheets and the export belongs to a site.
    assert all(n.startswith(("Mississauga", "Calgary")) for n in names[3:-1])
    # ...and the customer reports are named "<Warehouse> - <Report>".
    assert "Calgary - AMZ" in names and "Mississauga - Sysco" in names


def test_workbook_sheet_names_are_legal_and_unique():
    wb = load_workbook(io.BytesIO(build_workbook(_process(_multi_site()))))
    assert all(len(n) <= MAX_SHEET_NAME for n in wb.sheetnames)
    assert len({n.casefold() for n in wb.sheetnames}) == len(wb.sheetnames)


def test_long_warehouse_label_is_shortened_the_same_way_on_every_sheet():
    rules = load_ruleset()
    rules.site_display["CALGARY WAREHOUSE"] = "Calgary North Regional Distribution Centre"
    result = process_export(_multi_site(), AS_OF, ruleset=rules)

    wb = load_workbook(io.BytesIO(build_workbook(result)))
    assert all(len(n) <= MAX_SHEET_NAME for n in wb.sheetnames)
    assert len({n.casefold() for n in wb.sheetnames}) == len(wb.sheetnames)
    # One shortened label, reused across the whole block.
    block = [n for n in wb.sheetnames if n.startswith("Calgary North ")]
    assert {"Calgary North Issue Tracker", "Calgary North Not Started",
            "Calgary North Orders"} <= set(block)


def test_validation_warnings_sheet_is_written_last():
    result = _process(_rows({"ys": "Calgary Warehouse", "Sales Order Total": "TBD"}))
    assert not result.validation_errors.empty
    assert [n for n, _df, _o in sheet_plan(result)][-1] == VALIDATION_SHEET
    wb = load_workbook(io.BytesIO(build_workbook(result)))
    assert wb.sheetnames[-1] == VALIDATION_SHEET


# ---------------------------------------------------------------------------
# End-to-end validation against the prototype (skips if file absent)
# ---------------------------------------------------------------------------
@pytest.mark.skipif(not PROTOTYPE.exists(), reason="prototype workbook not available")
@pytest.mark.parametrize("report,rows,total", [
    ("AMZ", 12, 240636.72),
    ("PFG WHSE", 14, 256270.89),
    ("GFS", 10, None),
    ("Sysco", 11, None),
    ("Calgary CO-OP", 7, None),
    ("Pratts", 4, None),
])
def test_prototype_report_counts(report, rows, total):
    result = process_export(str(PROTOTYPE), AS_OF)
    df = result.reports[report]
    assert len(df) == rows, f"{report}: expected {rows} rows, got {len(df)}"
    if total is not None:
        got = float(pd.to_numeric(df["Sales Order Total"], errors="coerce").fillna(0).sum())
        assert abs(got - total) < 0.05, f"{report}: expected ${total}, got ${got}"


@pytest.mark.skipif(not PROTOTYPE.exists(), reason="prototype workbook not available")
def test_prototype_dedup_and_workbook_builds():
    result = process_export(str(PROTOTYPE), AS_OF)
    assert len(result.orders) == result.orders["LE Delivery"].nunique()
    xlsx = build_workbook(result)
    assert len(xlsx) > 10_000
    wb = load_workbook(io.BytesIO(xlsx))
    assert "Issue Tracker" in wb.sheetnames
